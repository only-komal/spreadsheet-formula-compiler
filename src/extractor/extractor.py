from openpyxl import load_workbook


def extract_cells(file_path):
    workbook = load_workbook(file_path, data_only=False)
    sheet = workbook.active

    cells = {}

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            if isinstance(cell.value, str) and cell.value.startswith("="):
                cells[cell.coordinate] = {
                    "type": "formula",
                    "formula": cell.value[1:]
                }
            else:
                cells[cell.coordinate] = {
                    "type": "value",
                    "value": cell.value
                }

    return cells
